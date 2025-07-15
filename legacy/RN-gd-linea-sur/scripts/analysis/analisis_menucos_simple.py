import openpyxl
from pathlib import Path

# Rutas a los archivos
data_path = Path("/Users/maxkeczeli/Proyects/estudio-gd-linea-sur/data")
archivos = {
    "Costo Los Menucos.xlsx": data_path / "Costo Los Menucos.xlsx",
    "Costo ganeración Los Menucos Nov24Ene25.xlsx": data_path / "Costo ganeración Los Menucos Nov24Ene25.xlsx",
    "LOS MENUCOS.xlsx": data_path / "LOS MENUCOS.xlsx"
}

print("=" * 80)
print("ANÁLISIS DE ARCHIVOS DE COSTOS - LOS MENUCOS")
print("=" * 80)

def analizar_excel(filepath, nombre):
    print(f"\n{'='*60}")
    print(f"ARCHIVO: {nombre}")
    print(f"{'='*60}")
    
    try:
        # Cargar el archivo
        wb = openpyxl.load_workbook(filepath, data_only=True)
        hojas = wb.sheetnames
        print(f"\nHojas disponibles: {hojas}")
        
        for hoja_nombre in hojas:
            print(f"\n{'-'*40}")
            print(f"HOJA: {hoja_nombre}")
            print(f"{'-'*40}")
            
            hoja = wb[hoja_nombre]
            
            # Obtener dimensiones
            max_row = hoja.max_row
            max_col = hoja.max_column
            print(f"Dimensiones: {max_row} filas x {max_col} columnas")
            
            # Leer primeras filas para entender la estructura
            print("\nPrimeras 10 filas:")
            for i in range(1, min(11, max_row + 1)):
                fila = []
                for j in range(1, min(10, max_col + 1)):
                    valor = hoja.cell(row=i, column=j).value
                    if valor is not None:
                        fila.append(str(valor)[:30])  # Limitar longitud para visualización
                if fila:
                    print(f"Fila {i}: {' | '.join(fila)}")
            
            # Buscar celdas con palabras clave
            keywords = ['costo', 'consumo', 'combustible', 'potencia', 'mw', 'kw', 
                       'gas', 'generac', 'motor', 'capex', 'opex', '$', 'litros', 'm3',
                       'total', 'mensual', 'diario', 'hora', 'precio']
            
            print("\nCeldas con información relevante:")
            encontradas = 0
            for row in range(1, min(50, max_row + 1)):  # Revisar primeras 50 filas
                for col in range(1, min(20, max_col + 1)):  # Revisar primeras 20 columnas
                    valor = hoja.cell(row=row, column=col).value
                    if valor and any(kw in str(valor).lower() for kw in keywords):
                        print(f"  [{row},{col}]: {str(valor)[:60]}")
                        encontradas += 1
                        if encontradas > 20:  # Limitar resultados
                            break
                if encontradas > 20:
                    break
            
        wb.close()
        
    except Exception as e:
        print(f"Error al leer archivo: {e}")

# Analizar cada archivo
for nombre, filepath in archivos.items():
    if filepath.exists():
        analizar_excel(filepath, nombre)
    else:
        print(f"\nARCHIVO NO ENCONTRADO: {nombre}")