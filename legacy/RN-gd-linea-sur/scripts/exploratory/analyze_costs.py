#!/usr/bin/env python3
import openpyxl
import os

def analyze_cost_file(file_path):
    """Analyze cost data from Excel file"""
    print(f"\nAnalyzing: {os.path.basename(file_path)}")
    print("=" * 60)
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        for sheet_name in wb.sheetnames:
            print(f"\nSheet: {sheet_name}")
            ws = wb[sheet_name]
            
            # Look for cost-related data
            for row in range(1, min(20, ws.max_row + 1)):
                row_data = []
                for col in range(1, min(10, ws.max_column + 1)):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value is not None:
                        row_data.append(str(cell_value)[:30])
                
                if row_data:
                    print(f"  Row {row}: {row_data}")
            
            # Look for specific cost indicators
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        value_lower = str(cell.value).lower()
                        if any(term in value_lower for term in ['costo', 'cost', 'usd', '$', 'precio', 'tarifa', 'kwh', 'mwh']):
                            # Print this row and the next few
                            print(f"\n  Found cost indicator at {cell.coordinate}: {cell.value}")
                            # Get surrounding cells
                            for i in range(max(1, cell.row - 1), min(ws.max_row + 1, cell.row + 3)):
                                row_values = []
                                for j in range(max(1, cell.column - 1), min(ws.max_column + 1, cell.column + 5)):
                                    val = ws.cell(row=i, column=j).value
                                    if val is not None:
                                        row_values.append(f"{val}")
                                if row_values:
                                    print(f"    Row {i}: {row_values}")
        
        wb.close()
        
    except Exception as e:
        print(f"  Error: {e}")

def main():
    print("=== LOS MENUCOS GENERATION COST ANALYSIS ===")
    
    cost_files = [
        '/Users/maxkeczeli/Proyects/estudio-gd-linea-sur/data/Costo Los Menucos.xlsx',
        '/Users/maxkeczeli/Proyects/estudio-gd-linea-sur/data/Costo ganeración Los Menucos Nov24Ene25.xlsx',
        '/Users/maxkeczeli/Proyects/estudio-gd-linea-sur/data/LOS MENUCOS.xlsx'
    ]
    
    for file_path in cost_files:
        if os.path.exists(file_path):
            analyze_cost_file(file_path)
    
    # Also check the summary file
    summary_file = '/Users/maxkeczeli/Proyects/estudio-gd-linea-sur/data/Datos Mayo 2024-Julio 2024 (Con Res 286 EPRE) (incluido gastos Los Menuc....xls'
    if os.path.exists(summary_file):
        print(f"\n\nChecking summary file for cost data...")
        try:
            import xlrd
            wb = xlrd.open_workbook(summary_file)
            print(f"Sheets: {wb.sheet_names()}")
            
            # Check first few sheets
            for sheet_name in wb.sheet_names()[:3]:
                sheet = wb.sheet_by_name(sheet_name)
                print(f"\nSheet: {sheet_name}")
                
                # Look for Los Menucos related data
                for row_idx in range(min(20, sheet.nrows)):
                    row_data = []
                    for col_idx in range(min(10, sheet.ncols)):
                        try:
                            cell_value = sheet.cell_value(row_idx, col_idx)
                            if cell_value:
                                cell_str = str(cell_value)
                                if 'menucos' in cell_str.lower() or 'costo' in cell_str.lower():
                                    print(f"  Row {row_idx}: Found relevant data")
                                    # Print this row
                                    for c in range(sheet.ncols):
                                        val = sheet.cell_value(row_idx, c)
                                        if val:
                                            print(f"    Col {c}: {val}")
                        except:
                            pass
                            
        except Exception as e:
            print(f"Error reading summary file: {e}")

if __name__ == "__main__":
    main()