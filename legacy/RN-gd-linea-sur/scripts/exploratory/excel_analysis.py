#!/usr/bin/env python3
import os
import glob
from datetime import datetime
import statistics

try:
    import pandas as pd
    import openpyxl
    import xlrd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    print("Warning: pandas not available, using basic analysis")

def read_excel_basic(file_path):
    """Read Excel file using openpyxl"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True)
        ws = wb.active
        
        data = []
        for i, row in enumerate(ws.iter_rows(min_row=7, values_only=True)):
            if i > 100:  # Limit rows for testing
                break
            if row[0] is not None:
                try:
                    record = {
                        'p_kw': float(row[0]) if row[0] else 0,
                        'q_kvar': float(row[1]) if row[1] else 0,
                        'v1': float(row[2]) if row[2] else 0,
                        'v2': float(row[3]) if row[3] else 0,
                        'v3': float(row[4]) if row[4] else 0,
                        'i1': float(row[5]) if row[5] else 0,
                        'i2': float(row[6]) if row[6] else 0,
                        'i3': float(row[7]) if row[7] else 0,
                        'date': row[8] if row[8] else ''
                    }
                    data.append(record)
                except:
                    pass
        
        wb.close()
        return data
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        return []

def analyze_station_data(data, station_name, nominal_voltage=33):
    """Analyze station data"""
    if not data:
        return None
    
    # Convert to proper units
    p_mw = [d['p_kw']/1000 for d in data]
    q_mvar = [d['q_kvar']/1000 for d in data]
    v_avg = [(d['v1'] + d['v2'] + d['v3'])/3/1000 for d in data]  # kV
    i_avg = [(d['i1'] + d['i2'] + d['i3'])/3 for d in data]  # A
    
    # Calculate statistics
    results = {
        'station': station_name,
        'records': len(data),
        'p_max_mw': max(p_mw) if p_mw else 0,
        'p_min_mw': min(p_mw) if p_mw else 0,
        'p_avg_mw': statistics.mean(p_mw) if p_mw else 0,
        'q_max_mvar': max(q_mvar) if q_mvar else 0,
        'q_min_mvar': min(q_mvar) if q_mvar else 0,
        'q_avg_mvar': statistics.mean(q_mvar) if q_mvar else 0,
        'v_max_kv': max(v_avg) if v_avg else 0,
        'v_min_kv': min(v_avg) if v_avg else 0,
        'v_avg_kv': statistics.mean(v_avg) if v_avg else 0,
        'v_min_pu': min(v_avg)/nominal_voltage if v_avg else 0,
        'v_max_pu': max(v_avg)/nominal_voltage if v_avg else 0,
        'i_max_a': max(i_avg) if i_avg else 0,
        'i_avg_a': statistics.mean(i_avg) if i_avg else 0,
    }
    
    # Calculate power factor
    s_mva = [((p**2 + q**2)**0.5) for p, q in zip(p_mw, q_mvar)]
    fp = [p/s if s > 0 else 0 for p, s in zip(p_mw, s_mva)]
    results['fp_min'] = min(fp) if fp else 0
    results['fp_avg'] = statistics.mean(fp) if fp else 0
    
    # Count voltage violations
    v_below_95 = sum(1 for v in v_avg if v < 0.95 * nominal_voltage)
    v_below_90 = sum(1 for v in v_avg if v < 0.90 * nominal_voltage)
    results['hours_below_95pu'] = v_below_95 * 0.25  # 15 min intervals
    results['hours_below_90pu'] = v_below_90 * 0.25
    
    return results

def process_all_stations():
    """Process all station data"""
    base_path = '/Users/maxkeczeli/Proyects/estudio-gd-linea-sur/data/Registros Línea Sur'
    
    # Define stations to analyze
    stations = [
        ('Pilcaniyeu_33kV', 'Pilcaniyeu2025/0425/ET4PI_33 042025.xlsx', 33),
        ('Jacobacci_Norte', 'Jacobacci2025/0425/ET2IJ_NORTE 0425.xlsx', 33),
        ('Jacobacci_Sur', 'Jacobacci2025/0425/ET2IJ_SUR 0425.xlsx', 33),
        ('Los_Menucos', 'Menucos2025/0425/ET2LM 0425.xlsx', 33),
        ('Maquinchao', 'Maquinchao2025/0425/ET2MA 0425.xlsx', 33)
    ]
    
    results = []
    
    for station_name, rel_path, nominal_v in stations:
        file_path = os.path.join(base_path, rel_path)
        
        if os.path.exists(file_path):
            print(f"\nAnalyzing {station_name}...")
            
            if PANDAS_AVAILABLE:
                try:
                    # Try pandas first
                    df = pd.read_excel(file_path, skiprows=5)
                    df.columns = ['P_kW', 'Q_kVAr', 'V1', 'V2', 'V3', 'I1', 'I2', 'I3', 'Date', 'Desc']
                    
                    # Convert to dict format for analysis
                    data = df.to_dict('records')
                    data = [{
                        'p_kw': d['P_kW'],
                        'q_kvar': d['Q_kVAr'],
                        'v1': d['V1'],
                        'v2': d['V2'],
                        'v3': d['V3'],
                        'i1': d['I1'],
                        'i2': d['I2'],
                        'i3': d['I3'],
                        'date': d['Date']
                    } for d in data[:100]]  # Limit for testing
                    
                    print(f"  Loaded {len(data)} records using pandas")
                except:
                    # Fallback to openpyxl
                    data = read_excel_basic(file_path)
                    print(f"  Loaded {len(data)} records using openpyxl")
            else:
                data = read_excel_basic(file_path)
                print(f"  Loaded {len(data)} records using openpyxl")
            
            # Analyze data
            analysis = analyze_station_data(data, station_name, nominal_v)
            if analysis:
                results.append(analysis)
                
                # Print key findings
                print(f"  Power: {analysis['p_min_mw']:.2f} - {analysis['p_max_mw']:.2f} MW (avg: {analysis['p_avg_mw']:.2f})")
                print(f"  Voltage: {analysis['v_min_kv']:.1f} - {analysis['v_max_kv']:.1f} kV ({analysis['v_min_pu']:.3f} - {analysis['v_max_pu']:.3f} pu)")
                print(f"  Power Factor: {analysis['fp_min']:.3f} - {analysis['fp_avg']:.3f}")
                print(f"  Hours below 0.95pu: {analysis['hours_below_95pu']:.1f}")
        else:
            print(f"\nFile not found: {file_path}")
    
    return results

def calculate_line_losses(results):
    """Estimate line losses between stations"""
    print("\n=== ESTIMATED LINE LOSSES ===")
    
    # Define line sections (approximate)
    sections = [
        ('Pilcaniyeu -> Jacobacci', 150, 'Pilcaniyeu_33kV', 'Jacobacci_Norte'),
        ('Jacobacci -> Maquinchao', 70, 'Jacobacci_Sur', 'Maquinchao'),
        ('Maquinchao -> Los Menucos', 50, 'Maquinchao', 'Los_Menucos')
    ]
    
    # Find station data
    station_data = {r['station']: r for r in results}
    
    for section_name, length_km, from_st, to_st in sections:
        if from_st in station_data and to_st in station_data:
            p_from = station_data[from_st]['p_avg_mw']
            p_to = station_data[to_st]['p_avg_mw']
            
            # Estimate losses (simplified)
            p_loss = p_from - p_to
            if p_loss > 0:
                loss_percent = (p_loss / p_from) * 100
                print(f"{section_name} ({length_km} km):")
                print(f"  Power in: {p_from:.2f} MW, Power out: {p_to:.2f} MW")
                print(f"  Losses: {p_loss:.2f} MW ({loss_percent:.1f}%)")
                print(f"  Loss per km: {(p_loss/length_km)*1000:.1f} kW/km")

def identify_critical_points(results):
    """Identify critical points for DG placement"""
    print("\n=== CRITICAL POINTS FOR DISTRIBUTED GENERATION ===")
    
    critical = []
    for r in results:
        score = 0
        reasons = []
        
        # Check voltage violations
        if r['v_min_pu'] < 0.95:
            score += 3
            reasons.append(f"Low voltage ({r['v_min_pu']:.3f} pu)")
        
        if r['hours_below_95pu'] > 10:
            score += 2
            reasons.append(f"Frequent voltage violations ({r['hours_below_95pu']:.0f} hours)")
        
        # Check reactive power
        if abs(r['q_avg_mvar']) > 0.3 * r['p_avg_mw']:
            score += 2
            reasons.append(f"High reactive demand ({r['q_avg_mvar']:.2f} MVAr)")
        
        # Check power factor
        if r['fp_avg'] < 0.95:
            score += 1
            reasons.append(f"Low power factor ({r['fp_avg']:.3f})")
        
        if score > 0:
            critical.append({
                'station': r['station'],
                'score': score,
                'reasons': reasons,
                'p_avg_mw': r['p_avg_mw'],
                'v_min_pu': r['v_min_pu']
            })
    
    # Sort by criticality
    critical.sort(key=lambda x: x['score'], reverse=True)
    
    for i, c in enumerate(critical, 1):
        print(f"\n{i}. {c['station']} (Score: {c['score']}/8)")
        print(f"   Average demand: {c['p_avg_mw']:.2f} MW")
        print(f"   Issues:")
        for reason in c['reasons']:
            print(f"   - {reason}")

def main():
    print("=== LÍNEA SUR - DISTRIBUTED GENERATION ANALYSIS ===")
    print(f"Analysis Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    
    # Process all stations
    results = process_all_stations()
    
    if results:
        # Summary table
        print("\n=== STATION SUMMARY ===")
        print(f"{'Station':<20} {'P_avg (MW)':<12} {'V_min (pu)':<12} {'FP_avg':<10} {'Hours<0.95pu':<15}")
        print("-" * 80)
        
        for r in results:
            print(f"{r['station']:<20} {r['p_avg_mw']:<12.2f} {r['v_min_pu']:<12.3f} {r['fp_avg']:<10.3f} {r['hours_below_95pu']:<15.1f}")
        
        # Calculate losses
        calculate_line_losses(results)
        
        # Identify critical points
        identify_critical_points(results)
        
        # Save results
        print("\n=== SAVING RESULTS ===")
        with open('station_analysis_results.txt', 'w') as f:
            f.write("Station Analysis Results\n")
            f.write("=" * 50 + "\n\n")
            for r in results:
                f.write(f"{r['station']}:\n")
                for k, v in r.items():
                    if k != 'station':
                        f.write(f"  {k}: {v}\n")
                f.write("\n")
        
        print("Results saved to station_analysis_results.txt")

if __name__ == "__main__":
    main()