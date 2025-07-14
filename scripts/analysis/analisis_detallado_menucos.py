import openpyxl
from pathlib import Path
from datetime import datetime

# Rutas a los archivos
data_path = Path("/Users/maxkeczeli/Proyects/estudio-gd-linea-sur/data")

def extraer_datos_costo_menucos():
    """Extrae datos del archivo Costo Los Menucos.xlsx"""
    filepath = data_path / "Costo Los Menucos.xlsx"
    wb = openpyxl.load_workbook(filepath, data_only=True)
    hoja = wb['Costos']
    
    print("\n" + "="*80)
    print("ANÁLISIS DETALLADO: Costo Los Menucos.xlsx")
    print("="*80)
    
    # Extraer datos de alquileres y costos
    alquileres = []
    combustible = []
    otros_costos = []
    
    for row in range(3, hoja.max_row + 1):
        concepto = hoja.cell(row=row, column=1).value
        if concepto:
            proveedor = hoja.cell(row=row, column=2).value
            factura = hoja.cell(row=row, column=3).value
            fecha = hoja.cell(row=row, column=4).value
            cantidad = hoja.cell(row=row, column=5).value
            unidad = hoja.cell(row=row, column=6).value
            moneda = hoja.cell(row=row, column=7).value
            importe_unit = hoja.cell(row=row, column=8).value
            importe_sin_iva = hoja.cell(row=row, column=9).value
            tipo_cambio = hoja.cell(row=row, column=10).value
            total_pesos = hoja.cell(row=row, column=11).value
            kw = hoja.cell(row=row, column=12).value
            
            registro = {
                'concepto': concepto,
                'fecha': fecha,
                'cantidad': cantidad,
                'unidad': unidad,
                'moneda': moneda,
                'importe_unit': importe_unit,
                'importe_sin_iva': importe_sin_iva,
                'tipo_cambio': tipo_cambio,
                'total_pesos': total_pesos,
                'kw': kw
            }
            
            concepto_lower = str(concepto).lower()
            if 'alquiler' in concepto_lower:
                alquileres.append(registro)
            elif 'gasoil' in concepto_lower or 'combustible' in concepto_lower or 'gas' in concepto_lower:
                combustible.append(registro)
            else:
                otros_costos.append(registro)
    
    # Resumen de alquileres
    print("\nALQUILERES DE GRUPOS GENERADORES:")
    print("-" * 60)
    equipos_alquiler = {}
    for alq in alquileres:
        equipo = alq['concepto'].split('kVA')[0] if 'kVA' in alq['concepto'] else alq['concepto'][:50]
        if equipo not in equipos_alquiler:
            equipos_alquiler[equipo] = {
                'cantidad_registros': 0,
                'total_dias': 0,
                'total_usd': 0,
                'kw': alq.get('kw', 0)
            }
        
        if alq['unidad'] == 'Día' and alq['cantidad']:
            equipos_alquiler[equipo]['cantidad_registros'] += 1
            equipos_alquiler[equipo]['total_dias'] += alq['cantidad'] or 0
            equipos_alquiler[equipo]['total_usd'] += alq['importe_sin_iva'] or 0
    
    for equipo, datos in equipos_alquiler.items():
        if datos['total_dias'] > 0:
            print(f"\nEquipo: {equipo}")
            print(f"  - Potencia: {datos['kw']} kW")
            print(f"  - Total días alquilados: {datos['total_dias']}")
            print(f"  - Total USD: ${datos['total_usd']:,.2f}")
            print(f"  - Costo promedio diario: ${datos['total_usd']/datos['total_dias']:,.2f} USD/día")
    
    # Resumen de combustible
    print("\n\nCOMBUSTIBLE:")
    print("-" * 60)
    total_litros = 0
    total_costo_combustible = 0
    for comb in combustible:
        if comb['unidad'] == 'Lts' and comb['cantidad']:
            total_litros += comb['cantidad'] or 0
            total_costo_combustible += comb['total_pesos'] or 0
            print(f"  - {comb['fecha']}: {comb['cantidad']:,.0f} litros, ${comb['total_pesos']:,.2f}")
    
    if total_litros > 0:
        print(f"\nTotal combustible: {total_litros:,.0f} litros")
        print(f"Total costo: ${total_costo_combustible:,.2f}")
        print(f"Precio promedio: ${total_costo_combustible/total_litros:,.2f}/litro")
    
    # Otros costos
    print("\n\nOTROS COSTOS:")
    print("-" * 60)
    for otro in otros_costos[:10]:  # Mostrar primeros 10
        print(f"  - {otro['concepto']}: ${otro['total_pesos']:,.2f} ({otro['unidad']})")
    
    wb.close()
    return alquileres, combustible, otros_costos

def extraer_datos_generacion_nov_ene():
    """Extrae datos del archivo Costo ganeración Los Menucos Nov24Ene25.xlsx"""
    filepath = data_path / "Costo ganeración Los Menucos Nov24Ene25.xlsx"
    wb = openpyxl.load_workbook(filepath, data_only=True)
    hoja = wb['Alquileres']
    
    print("\n\n" + "="*80)
    print("ANÁLISIS DETALLADO: Costo ganeración Los Menucos Nov24Ene25.xlsx")
    print("="*80)
    
    # Extraer datos por mes
    datos_mensuales = {}
    mes_actual = None
    
    for row in range(2, hoja.max_row + 1):
        # Detectar cambio de mes
        celda_mes = hoja.cell(row=row, column=1).value
        if celda_mes and ('2024' in str(celda_mes) or '2025' in str(celda_mes)):
            mes_actual = celda_mes
            datos_mensuales[mes_actual] = []
            continue
        
        # Extraer datos si hay concepto
        concepto = hoja.cell(row=row, column=1).value
        if concepto and mes_actual and concepto not in ['Concepto', None]:
            registro = {
                'mes': mes_actual,
                'concepto': concepto,
                'cantidad': hoja.cell(row=row, column=5).value,
                'unidad': hoja.cell(row=row, column=6).value,
                'importe_unit_usd': hoja.cell(row=row, column=8).value,
                'importe_sin_iva_usd': hoja.cell(row=row, column=9).value,
                'tipo_cambio': hoja.cell(row=row, column=10).value,
                'total_pesos': hoja.cell(row=row, column=11).value,
                'kw': hoja.cell(row=row, column=12).value
            }
            datos_mensuales[mes_actual].append(registro)
    
    # Mostrar resumen por mes
    for mes, registros in datos_mensuales.items():
        print(f"\n{mes}:")
        print("-" * 40)
        total_mes_usd = 0
        total_mes_pesos = 0
        
        for reg in registros:
            print(f"  - {reg['concepto'][:50]}")
            print(f"    Cantidad: {reg['cantidad']} {reg['unidad']}")
            print(f"    Importe: ${reg['importe_sin_iva_usd']:,.2f} USD")
            if reg['total_pesos']:
                print(f"    Total pesos: ${reg['total_pesos']:,.2f}")
            if reg['kw']:
                print(f"    Potencia: {reg['kw']} kW")
            
            total_mes_usd += reg['importe_sin_iva_usd'] or 0
            total_mes_pesos += reg['total_pesos'] or 0
        
        print(f"\n  TOTAL MES: ${total_mes_usd:,.2f} USD / ${total_mes_pesos:,.2f} ARS")
    
    wb.close()
    return datos_mensuales

def extraer_datos_los_menucos():
    """Extrae datos del archivo LOS MENUCOS.xlsx"""
    filepath = data_path / "LOS MENUCOS.xlsx"
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    print("\n\n" + "="*80)
    print("ANÁLISIS DETALLADO: LOS MENUCOS.xlsx")
    print("="*80)
    
    # Analizar Hoja1 - Comparativa de abastecimiento
    hoja1 = wb['Hoja1']
    print("\nHOJA 1: COMPARATIVA DE ABASTECIMIENTO")
    print("-" * 60)
    
    # Extraer datos de cargo fijo
    print("\nCARGO FIJO (ALQUILER):")
    for row in range(10, 15):
        proveedor = hoja1.cell(row=row, column=1).value
        factura = hoja1.cell(row=row, column=2).value
        fecha = hoja1.cell(row=row, column=3).value
        cantidad = hoja1.cell(row=row, column=4).value
        
        if proveedor and proveedor != 'Proveedor':
            print(f"  Proveedor: {proveedor}")
            print(f"  Fecha: {fecha}")
            print(f"  Cantidad: {cantidad} días")
            
            # Buscar importes
            for col in range(5, 20):
                valor = hoja1.cell(row=row, column=col).value
                if valor and isinstance(valor, (int, float)) and valor > 100:
                    unidad = hoja1.cell(row=8, column=col).value or hoja1.cell(row=9, column=col).value
                    print(f"  Importe: {valor:,.2f} {unidad}")
    
    # Extraer datos de combustible
    print("\n\nCOMBUSTIBLE Y CONSUMOS:")
    # Buscar fila de consumos específicos
    for row in range(15, 25):
        celda = hoja1.cell(row=row, column=4).value
        if celda:
            if 'consumo' in str(celda).lower():
                print(f"  {celda}: {hoja1.cell(row=row, column=5).value}")
            elif 'equipo' in str(celda).lower():
                valor = hoja1.cell(row=row, column=5).value
                unidad = hoja1.cell(row=row, column=6).value
                if valor:
                    print(f"  {celda}: {valor} {unidad if unidad else ''}")
    
    # Buscar datos de demanda
    print("\n\nDEMANDA Y COSTOS:")
    for row in range(20, 30):
        celda = hoja1.cell(row=row, column=4).value
        if celda and 'demanda' in str(celda).lower():
            for col in range(5, 15):
                valor = hoja1.cell(row=row, column=col).value
                encabezado = hoja1.cell(row=22, column=col).value
                if valor and encabezado:
                    print(f"  {encabezado}: {valor}")
    
    # Analizar Hoja3 - Análisis económico
    hoja3 = wb['Hoja3']
    print("\n\nHOJA 3: ANÁLISIS ECONÓMICO")
    print("-" * 60)
    
    # Extraer parámetros económicos
    parametros = {}
    for row in range(1, 12):
        param = hoja3.cell(row=row, column=1).value or hoja3.cell(row=row, column=4).value
        valor = hoja3.cell(row=row, column=2).value or hoja3.cell(row=row, column=6).value
        unidad = hoja3.cell(row=row, column=3).value or hoja3.cell(row=row, column=7).value
        
        if param and valor:
            print(f"{param}: {valor} {unidad if unidad else ''}")
            parametros[str(param)] = {'valor': valor, 'unidad': unidad}
    
    # Extraer datos de CAPEX/OPEX
    print("\n\nCAPEX Y OPEX:")
    for row in range(12, 20):
        concepto = hoja3.cell(row=row, column=4).value
        if concepto and ('CAPEX' in str(concepto) or 'OPEX' in str(concepto) or 'Costo' in str(concepto)):
            for col in range(5, 14):
                valor = hoja3.cell(row=row, column=col).value
                if valor and isinstance(valor, (int, float)):
                    print(f"  {concepto}: ${valor:,.2f}")
                    break
    
    wb.close()
    return parametros

# Ejecutar análisis
print("INICIANDO ANÁLISIS DETALLADO DE ARCHIVOS DE LOS MENUCOS")
print("="*80)

# Análisis archivo 1
try:
    alquileres, combustible, otros = extraer_datos_costo_menucos()
except Exception as e:
    print(f"Error analizando Costo Los Menucos.xlsx: {e}")

# Análisis archivo 2
try:
    datos_mensuales = extraer_datos_generacion_nov_ene()
except Exception as e:
    print(f"Error analizando Costo ganeración Los Menucos Nov24Ene25.xlsx: {e}")

# Análisis archivo 3
try:
    parametros_economicos = extraer_datos_los_menucos()
except Exception as e:
    print(f"Error analizando LOS MENUCOS.xlsx: {e}")

print("\n\n" + "="*80)
print("ANÁLISIS COMPLETADO")
print("="*80)