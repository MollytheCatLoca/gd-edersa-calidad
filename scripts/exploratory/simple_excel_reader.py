#!/usr/bin/env python3
from openpyxl import load_workbook
import statistics
from datetime import datetime

def read_measurement_file(file_path, max_rows=100):
    """Read measurement data from Excel file"""
    print(f"\nReading: {file_path}")
    
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        
        # Find data start row by looking for numeric data
        data_start_row = None
        for row_idx in range(1, 20):  # Check first 20 rows
            row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
            if row[0] is not None and isinstance(row[0], (int, float)):
                data_start_row = row_idx
                break
        
        if not data_start_row:
            print("Could not find data start row")
            return []
        
        print(f"Data starts at row {data_start_row}")
        
        # Read data
        data = []
        row_count = 0
        
        for row in ws.iter_rows(min_row=data_start_row, values_only=True):
            if row_count >= max_rows:
                break
                
            if row[0] is not None:
                try:
                    record = {
                        'p_mw': float(row[0]) / 1000 if row[0] else 0,  # kW to MW
                        'q_mvar': float(row[1]) / 1000 if row[1] else 0,  # kVAr to MVAr
                        'v1_kv': float(row[2]) / 1000 if row[2] else 0,  # V to kV
                        'v2_kv': float(row[3]) / 1000 if row[3] else 0,
                        'v3_kv': float(row[4]) / 1000 if row[4] else 0,
                        'i1_a': float(row[5]) if row[5] else 0,
                        'i2_a': float(row[6]) if row[6] else 0,
                        'i3_a': float(row[7]) if row[7] else 0,
                        'date': str(row[8]) if row[8] else ''
                    }
                    
                    # Calculate averages
                    record['v_avg_kv'] = (record['v1_kv'] + record['v2_kv'] + record['v3_kv']) / 3
                    record['i_avg_a'] = (record['i1_a'] + record['i2_a'] + record['i3_a']) / 3
                    
                    # Calculate apparent power and power factor
                    s_mva = (record['p_mw']**2 + record['q_mvar']**2)**0.5
                    record['s_mva'] = s_mva
                    record['fp'] = record['p_mw'] / s_mva if s_mva > 0 else 0
                    
                    data.append(record)
                    row_count += 1
                    
                    # Show first few records
                    if row_count <= 5:
                        print(f"  {record['date']}: P={record['p_mw']:.2f} MW, V={record['v_avg_kv']:.1f} kV, FP={record['fp']:.3f}")
                        
                except Exception as e:
                    if row_count < 5:
                        print(f"  Error in row: {e}")
        
        wb.close()
        print(f"  Total records read: {len(data)}")
        return data
        
    except Exception as e:
        print(f"  Error reading file: {e}")
        return []

def analyze_station(data, station_name, nominal_kv=33):
    """Analyze station data and return key metrics"""
    if not data:
        return None
    
    # Extract values for analysis
    p_values = [d['p_mw'] for d in data]
    q_values = [d['q_mvar'] for d in data]
    v_values = [d['v_avg_kv'] for d in data]
    fp_values = [d['fp'] for d in data]
    
    results = {
        'station': station_name,
        'records': len(data),
        'p_max_mw': max(p_values),
        'p_min_mw': min(p_values),
        'p_avg_mw': statistics.mean(p_values),
        'q_max_mvar': max(q_values),
        'q_min_mvar': min(q_values),
        'q_avg_mvar': statistics.mean(q_values),
        'v_max_kv': max(v_values),
        'v_min_kv': min(v_values),
        'v_avg_kv': statistics.mean(v_values),
        'v_min_pu': min(v_values) / nominal_kv,
        'v_avg_pu': statistics.mean(v_values) / nominal_kv,
        'fp_min': min(fp_values),
        'fp_avg': statistics.mean(fp_values),
        'hours_below_095': sum(1 for v in v_values if v < 0.95 * nominal_kv) * 0.25,
        'hours_below_090': sum(1 for v in v_values if v < 0.90 * nominal_kv) * 0.25
    }
    
    return results

def main():
    print("=== LÍNEA SUR - ELECTRICAL MEASUREMENTS ANALYSIS ===")
    print(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n")
    
    # Define files to analyze
    base_path = '/Users/maxkeczeli/Proyects/estudio-gd-linea-sur/data/Registros Línea Sur'
    
    stations = [
        # Recent data (April 2025)
        ('Pilcaniyeu 33kV', f'{base_path}/Pilcaniyeu2025/0425/ET4PI_33 042025.xlsx', 33),
        ('Jacobacci Norte', f'{base_path}/Jacobacci2025/0425/ET2IJ_NORTE 0425.xlsx', 33),
        ('Jacobacci Sur', f'{base_path}/Jacobacci2025/0425/ET2IJ_SUR 0425.xlsx', 33),
        ('Los Menucos', f'{base_path}/Menucos2025/0425/ET2LM 0425.xlsx', 33),
        ('Maquinchao', f'{base_path}/Maquinchao2025/0425/ET2MA 0425.xlsx', 33),
        
        # Also try October 2024 data (xls format)
        ('Pilcaniyeu 33kV Oct24', f'{base_path}/Pilcaniyeu2024/1024/ET4PI_33 1024.xls', 33),
        ('Los Menucos Oct24', f'{base_path}/Menucos2024/1024/ET2LM 1024.xls', 33),
    ]
    
    all_results = []
    
    for station_name, file_path, nominal_kv in stations:
        data = read_measurement_file(file_path, max_rows=200)
        
        if data:
            results = analyze_station(data, station_name, nominal_kv)
            if results:
                all_results.append(results)
    
    # Display results summary
    if all_results:
        print("\n\n=== ANALYSIS SUMMARY ===")
        print(f"{'Station':<25} {'Records':<10} {'P_avg(MW)':<12} {'V_min(kV)':<12} {'V_min(pu)':<12} {'FP_avg':<10}")
        print("-" * 90)
        
        for r in all_results:
            print(f"{r['station']:<25} {r['records']:<10} {r['p_avg_mw']:<12.2f} {r['v_min_kv']:<12.1f} {r['v_min_pu']:<12.3f} {r['fp_avg']:<10.3f}")
        
        # Identify critical stations
        print("\n\n=== CRITICAL STATIONS (V < 0.95 pu) ===")
        critical = [r for r in all_results if r['v_min_pu'] < 0.95]
        
        for r in sorted(critical, key=lambda x: x['v_min_pu']):
            print(f"\n{r['station']}:")
            print(f"  Minimum voltage: {r['v_min_kv']:.1f} kV ({r['v_min_pu']:.3f} pu)")
            print(f"  Average voltage: {r['v_avg_kv']:.1f} kV ({r['v_avg_pu']:.3f} pu)")
            print(f"  Hours below 0.95 pu: {r['hours_below_095']:.0f}")
            print(f"  Average demand: {r['p_avg_mw']:.2f} MW")
            print(f"  Average power factor: {r['fp_avg']:.3f}")
        
        # Save detailed results
        with open('measurement_analysis_results.txt', 'w') as f:
            f.write("LÍNEA SUR - MEASUREMENT ANALYSIS RESULTS\n")
            f.write("=" * 50 + "\n\n")
            
            for r in all_results:
                f.write(f"{r['station']}:\n")
                for key, value in r.items():
                    if key != 'station':
                        f.write(f"  {key}: {value}\n")
                f.write("\n")
        
        print("\n\nDetailed results saved to: measurement_analysis_results.txt")
    else:
        print("\nNo data could be analyzed. Please check file formats.")

if __name__ == "__main__":
    main()